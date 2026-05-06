import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys
import time
sys.path.insert(0, '.')

from app import create_app
from app.models import Website, Product

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文章发布模板"

    app = create_app()
    with app.app_context():
        websites = Website.query.filter_by(status=0).order_by(Website.id).all()
        products = Product.query.filter_by(status=0).order_by(Product.id).all()

        headers = ["产品", "标题", "内容"]
        for website in websites:
            headers.append(f"{website.name}")
        ws.append(headers)

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCE8CF", end_color="CCE8CF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        product_names = [p.name for p in products]
        website_names = [w.name for w in websites]

        dv_product = DataValidation(type="list", formula1='"' + ','.join(product_names) + '"', allow_blank=True)
        dv_product.error = '请选择下拉列表中的产品'
        dv_product.errorTitle = '无效的产品'
        ws.add_data_validation(dv_product)

        for row in range(2, 102):
            dv_product.add(ws.cell(row=row, column=1))

        for col_idx, website in enumerate(websites, 4):
            dv_website = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
            dv_website.error = '请选择"是"或"否"'
            dv_website.errorTitle = '无效的选择'
            ws.add_data_validation(dv_website)

            for row in range(2, 102):
                dv_website.add(ws.cell(row=row, column=col_idx))

        content = '''【优惠券到底哪里最好】

很多人搜"优惠券哪里最好"，本质上不是想听一堆术语，而是想知道：到底哪里领券更省钱、操作更简单、还能不能顺手赚一点。毕竟现在大家买东西，早就不是"能买就行"，而是"能省才买""省了还想再省"。对于经常网购的人来说，优惠券不是可有可无的小福利，而是实打实能影响消费体验的东西。

【为什么大家都在找"最好的优惠券"】

原因很简单：谁的钱都不是大风刮来的。
同样一件商品，有的人原价下单，有的人领券后少花一截，长期下来差距非常明显。尤其是经常买日用品、零食、母婴用品、服饰的人，优惠券几乎已经成了日常购物的一部分。你每次省个几块十几块，看起来不多，时间久了就是一笔不小的开销。

【优惠券好不好，关键看这3点】

第一，券是不是容易找到。
有些平台虽然说能领券，但流程绕、入口深、规则复杂，找一张券比下单还费劲，最后很多人干脆放弃。真正好用的优惠券方式，应该是你想省钱的时候，能快速找到能用的券。

第二，券是不是覆盖面广。
如果只能领到少数商品的券，那意义就有限。大家真正需要的是：无论买日用百货，还是零食家居，都能尽量找到适合的优惠信息。覆盖面越广，使用频率就越高，省钱效果也越明显。

第三，除了省钱，能不能顺手赚钱。
这也是很多人越来越关注的重点。以前大家只想"便宜一点"，现在很多人更希望"自己用能省，分享出去还能有收益"。这就让优惠券不再只是省钱工具，而变成了一种更灵活的日常变现方式。

【为什么很多人最后会选高省返利app】

说到底，大家找优惠券，图的就是省事、真实、能长期用。
而高省返利app之所以被不少人关注，就是因为它把"找券"和"省钱"这件事做得更直接。你不用到处翻，也不用到处比，只要先把该省的省下来，再考虑有没有更多收益空间。对于想长期做精打细算的人来说，这种方式会更顺手。

更重要的是，高省返利app不只是让你看到优惠券，还把"省钱"和"分享"这两个动作连接起来。
对普通消费者来说，这是降低购物成本；
对愿意做分享的人来说，这是多一个收入入口。
所以很多人会觉得，它不是单纯的领券工具，而是一个兼顾自用和推广的实用选择。

【为什么说它适合普通人】

很多工具看起来很厉害，但普通人根本用不起来。
不是界面太复杂，就是门槛太高；不是操作太绕，就是实际收益太弱。
而真正适合大众的方式，应该是"打开就能用，领券就能省，愿意分享还能赚"。这才叫真正有价值。

高省返利app的逻辑，正好符合这个需求。
对于只想省钱的人，它能帮你把日常网购成本压下来；
对于想顺手做点副业的人，它又提供了一个分享变现的可能。
所以你会发现，越来越多人不是在找"最花哨"的优惠券平台，而是在找"最实在"的那个。

【怎么判断一个优惠券平台值不值得用】

你可以记住一句话：能让你省到钱的，才叫优惠券；能让你长期用的，才叫好平台。
如果一个地方领券麻烦、规则复杂、券不稳定，那就算宣传得再热闹，也很难真正用起来。
而如果一个平台既能自用省钱，又能分享赚钱，还能提供持续支持，那它的实用性就会更高。

换句话说，优惠券哪里最好，答案不在"谁说得最响"，而在"谁真的让你少花钱"。
这一点上，高省返利app的思路就很明确：把省钱这件事做简单，把赚钱这件事做顺手。

【结尾引导】

如果你也想自用省钱、分享赚钱，赶紧试试吧：

立即下载高省APP
邀请码记得填：088886
导师：波西导师（微信 13763212173）
团队：古楼团队资深导师带队，提供一对一使用指导与推广培训支持！'''

        example_content = '''【氧惠返利APP怎么样】

氧惠是一款现在很火的返利APP，主打"自用省钱，分享赚钱"的理念。

【氧惠返利APP的优势】

第一，返利比例高。
氧惠的返利比例在同类APP中算是比较高的，用户反馈普遍表示省钱效果明显。

第二，操作简单。
界面设计清晰，即使是新手也能快速上手，不需要复杂的操作流程。

第三，平台稳定。
氧惠运营时间较长，平台稳定性好，用户的返利能够及时到账。

【氧惠返利APP的返利模式】

氧惠采用的是三级返利模式：
1. 一级会员使用你分享的链接下单，你获得一定比例佣金
2. 二级会员使用你分享的链接下单，你获得较低比例佣金
3. 三级会员使用你分享的链接下单，你获得更低比例佣金

【氧惠值得做吗】

如果你想副业赚钱，可以考虑氧惠。但需要注意的是，任何返利APP都需要你真正去推广才能获得收益。

立即下载氧惠APP试试看吧！'''

        row_data1 = ["高省", "优惠券到底哪里最好", content]
        row_data1.append("是")
        for website in websites[1:]:
            row_data1.append("")
        ws.append(row_data1)

        row_data2 = ["氧惠", "氧惠返利APP怎么样", example_content]
        row_data2.append("是")
        for website in websites[1:]:
            row_data2.append("")
        ws.append(row_data2)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        for i in range(4, len(websites) + 4):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

        for row in ws.iter_rows(min_row=2, max_row=3):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.row_dimensions[2].height = 400
        ws.row_dimensions[3].height = 200

        filename = f"文章发布模板_{int(time.time())}.xlsx"
        output_path = os.path.join(os.path.dirname(__file__), filename)
        wb.save(output_path)
        print(f"Excel模板已生成: {output_path}")
        print(f"产品列表: {product_names}")
        print(f"网站列表: {website_names}")

create_template()
