import openpyxl

# 打开现有的Excel模板
wb = openpyxl.load_workbook('文章发布模板.xlsx')
ws = wb.active

# 获取表头
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print('=== 现有Excel模板结构 ===')
print('表头字段:', headers)
print()

# 获取示例数据
print('示例数据:')
for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
    print(row)

# 打印字段说明
print()
print('=== 字段说明 ===')
for i, header in enumerate(headers):
    print(f'{i+1}. {header}')
