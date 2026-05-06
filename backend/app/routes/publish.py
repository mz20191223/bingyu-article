from flask import Blueprint, request, jsonify, g, send_file
from app.routes.auth import token_required
from app.models import db, PublishRecord, Website, Product
from app.services.publish_service import generate_article, publish_article
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import io

bp = Blueprint('publish', __name__)


@bp.route('/generate', methods=['POST'])
@token_required
def generate():
    data = request.get_json()
    product_id = data.get('productId')
    website_ids = data.get('websiteIds', [])
    content_template_id = data.get('contentTemplateId')
    title_template_id = data.get('titleTemplateId')
    model_id = data.get('modelId')
    keyword_ids = data.get('keywordIds', [])

    if not product_id:
        return jsonify({'code': 400, 'msg': '请选择产品', 'data': None})
    if not website_ids:
        return jsonify({'code': 400, 'msg': '请选择目标网站', 'data': None})

    try:
        result = generate_article(product_id, website_ids, content_template_id, title_template_id, model_id, keyword_ids)
        return jsonify({'code': 200, 'msg': '生成成功', 'data': result})
    except Exception as e:
        return jsonify({'code': 500, 'msg': f'生成失败: {str(e)}', 'data': None})


@bp.route('/submit', methods=['POST'])
@token_required
def submit():
    data = request.get_json()
    product_id = data.get('productId')
    website_ids = data.get('websiteIds', [])
    title = data.get('title')
    content = data.get('content')
    model_id = data.get('modelId')
    title_template_id = data.get('titleTemplateId')
    content_template_id = data.get('contentTemplateId')

    if not product_id:
        return jsonify({'code': 400, 'msg': '请选择产品', 'data': None})
    if not website_ids:
        return jsonify({'code': 400, 'msg': '请选择目标网站', 'data': None})
    if not title or not content:
        return jsonify({'code': 400, 'msg': '标题和内容不能为空', 'data': None})

    try:
        result = publish_article(product_id, website_ids, title, content, model_id, title_template_id, content_template_id, g.user_id)
        return jsonify({'code': 200, 'msg': '发布成功', 'data': result})
    except Exception as e:
        return jsonify({'code': 500, 'msg': f'发布失败: {str(e)}', 'data': None})


@bp.route('/templates/download', methods=['GET'])
@token_required
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文章发布模板"

    websites = Website.query.filter_by(status=0).order_by(Website.id).all()
    products = Product.query.filter_by(status=0).order_by(Product.id).all()

    headers = ["产品", "标题", "内容"]
    for website in websites:
        headers.append(f"{website.name}")
    ws.append(headers)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CCE8CF", end_color="CCE8CF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    product_names = [p.name for p in products]

    dv_product = DataValidation(type="list", formula1='"' + ','.join(product_names) + '"', allow_blank=True)
    dv_product.error = '请选择下拉列表中的产品'
    dv_product.errorTitle = '无效的产品'
    ws.add_data_validation(dv_product)

    for row in range(2, 102):
        dv_product.add(ws.cell(row=row, column=1))

    for col_idx in range(4, len(websites) + 4):
        dv_website = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
        dv_website.error = '请选择"是"或"否"'
        dv_website.errorTitle = '无效的选择'
        ws.add_data_validation(dv_website)

        for row in range(2, 102):
            dv_website.add(ws.cell(row=row, column=col_idx))

    example_content = '''【示例标题】

这是第一段内容，段落之间要用空行分隔。

这是第二段内容，图片会插入到这个位置。

这是第三段内容。

【结尾】

这是结尾内容。'''

    row_data = ["高省", "示例标题", example_content]
    for website in websites:
        if website.id == 1:
            row_data.append("是")
        else:
            row_data.append("")
    ws.append(row_data)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60
    for i in range(4, len(websites) + 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

    for row in ws.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[2].height = 200

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='文章发布模板.xlsx'
    )
