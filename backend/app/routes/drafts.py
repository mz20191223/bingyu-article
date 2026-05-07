from flask import Blueprint, request, jsonify, g, send_file
from app.routes.auth import token_required
from app.models import db, Draft, Product, Website
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import io
from datetime import datetime

bp = Blueprint('drafts', __name__)


@bp.route('/drafts', methods=['GET'])
@token_required
def get_drafts():
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 10, type=int)
    
    query = Draft.query.filter_by(status=0).order_by(Draft.update_time.desc())
    pagination = query.paginate(page=page, per_page=size, error_out=False)
    
    result = []
    for draft in pagination.items:
        result.append({
            'id': draft.id,
            'name': draft.name,
            'data': draft.data,
            'createTime': draft.create_time.strftime('%Y-%m-%d %H:%M:%S') if draft.create_time else None,
            'updateTime': draft.update_time.strftime('%Y-%m-%d %H:%M:%S') if draft.update_time else None
        })
    
    return jsonify({
        'code': 200,
        'msg': 'success',
        'data': {
            'list': result,
            'page': page,
            'pageSize': size,
            'total': pagination.total
        }
    })


@bp.route('/drafts/<int:id>', methods=['GET'])
@token_required
def get_draft(id):
    draft = Draft.query.get(id)
    if not draft or draft.status != 0:
        return jsonify({'code': 404, 'msg': '草稿不存在', 'data': None})
    
    return jsonify({
        'code': 200,
        'msg': 'success',
        'data': {
            'id': draft.id,
            'name': draft.name,
            'data': draft.data
        }
    })


@bp.route('/drafts', methods=['POST'])
@token_required
def save_draft():
    data = request.get_json()
    name = data.get('name')
    draft_data = data.get('data')
    
    if not name:
        return jsonify({'code': 400, 'msg': '草稿名称不能为空', 'data': None})
    if not draft_data:
        return jsonify({'code': 400, 'msg': '草稿数据不能为空', 'data': None})
    
    draft_id = data.get('id')
    if draft_id:
        draft = Draft.query.get(draft_id)
        if not draft:
            return jsonify({'code': 404, 'msg': '草稿不存在', 'data': None})
        draft.name = name
        draft.data = draft_data
    else:
        draft = Draft(
            name=name,
            data=draft_data,
            create_by=g.user_id
        )
        db.session.add(draft)
    
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '保存成功', 'data': {'id': draft.id}})


@bp.route('/drafts/<int:id>', methods=['DELETE'])
@token_required
def delete_draft(id):
    draft = Draft.query.get(id)
    if not draft:
        return jsonify({'code': 404, 'msg': '草稿不存在', 'data': None})
    
    draft.status = 1
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '删除成功', 'data': None})


@bp.route('/drafts/<int:id>/export', methods=['GET'])
@token_required
def export_draft(id):
    draft = Draft.query.get(id)
    if not draft or draft.status != 0:
        return jsonify({'code': 404, 'msg': '草稿不存在', 'data': None})
    
    try:
        import json
        draft_data = json.loads(draft.data)
    except:
        return jsonify({'code': 400, 'msg': '草稿数据格式错误', 'data': None})
    
    websites = Website.query.filter_by(status=0).order_by(Website.id).all()
    products = Product.query.filter_by(status=0).order_by(Product.id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文章发布模板"
    
    headers = ["产品", "标题", "内容"]
    website_name_map = {}
    for website in websites:
        headers.append(f"{website.name}")
        website_name_map[website.id] = website.name
    ws.append(headers)
    
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CCE8CF", end_color="CCE8CF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    product_names = [p.name for p in products]
    product_id_map = {p.id: p.name for p in products}
    
    dv_product = DataValidation(type="list", formula1='"' + ','.join(product_names) + '"', allow_blank=True)
    dv_product.error = '请选择下拉列表中的产品'
    dv_product.errorTitle = '无效的产品'
    ws.add_data_validation(dv_product)
    
    row_num = 2
    for row_data in draft_data:
        product_name = product_id_map.get(row_data.get('productId'), '')
        title = row_data.get('title', '')
        content = row_data.get('content', '')
        
        row = [product_name, title, content]
        
        website_ids = row_data.get('websiteIds', [])
        for website in websites:
            row.append("是" if website.id in website_ids else "")
        
        ws.append(row)
        
        dv_product.add(ws.cell(row=row_num, column=1))
        
        for col_idx in range(4, len(websites) + 4):
            dv_website = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
            dv_website.error = '请选择"是"或"否"'
            dv_website.errorTitle = '无效的选择'
            ws.add_data_validation(dv_website)
            dv_website.add(ws.cell(row=row_num, column=col_idx))
        
        row_num += 1
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60
    for i in range(4, len(websites) + 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    
    for row in ws.iter_rows(min_row=2, max_row=row_num - 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{draft.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
